"""
Geração de ficheiros Excel para a página Relatórios (Recepção, Caixa, Toxicológico).
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

MESES_PT = (
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
)


def titulo_mes_ano(ref: date) -> str:
    return f"{MESES_PT[ref.month - 1].upper()}/{ref.year}"


def _sim_nao_excel(v: object) -> str:
    if v is True or v == 1:
        return "Sim"
    if v is False or v == 0:
        return "Não"
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "Não"
    s = str(v).strip().casefold()
    if s in ("true", "t", "1", "sim", "s"):
        return "Sim"
    return "Não"


def _formatar_planilha(
    ws,
    *,
    titulo: str,
    ncol: int,
    linha_header: int = 2,
) -> None:
    azul_escuro = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    azul_claro = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    branco = Font(color="FFFFFF", bold=True)
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max(ncol, 1),
    )
    c1 = ws.cell(row=1, column=1, value=titulo)
    c1.font = Font(bold=True, size=13, color="FFFFFF")
    c1.fill = azul_escuro
    c1.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, ncol + 1):
        h = ws.cell(row=linha_header, column=col)
        h.font = branco
        h.fill = azul_claro
        h.alignment = Alignment(horizontal="center", vertical="center")


def montar_xlsx_generico(df: pd.DataFrame, titulo: str, sheet_name: str = "Relatório") -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, startrow=1, index=False)
        ws = writer.sheets[sheet_name]
        _formatar_planilha(ws, titulo=titulo, ncol=len(df.columns))
    buf.seek(0)
    return buf.getvalue()


def dataframe_recepcao_relatorio(df: pd.DataFrame) -> pd.DataFrame:
    out = df.rename(
        columns={
            "data_hora": "Data",
            "nome": "Nome",
            "tipo": "Tipo",
            "empresa": "Empresa",
            "exames": "Exames",
            "valor": "Valor",
            "pagamento": "Pagamento",
            "telefone": "Telefone",
            "cpf": "CPF",
        }
    )
    return out


def dataframe_caixa_relatorio(df: pd.DataFrame) -> pd.DataFrame:
    linhas = []
    for _, r in df.iterrows():
        tipo = str(r.get("tipo", "")).strip().casefold()
        val = r.get("valor")
        try:
            v = float(val) if val is not None and not (isinstance(val, float) and pd.isna(val)) else None
        except (TypeError, ValueError):
            v = None
        entrada = ""
        saida = ""
        if "entrada" in tipo:
            entrada = v if v is not None else ""
        elif "said" in tipo or tipo == "saída":
            saida = v if v is not None else ""
        linhas.append(
            {
                "Data": r.get("data_mov"),
                "Descrição": r.get("descricao"),
                "Entrada": entrada,
                "Saída": saida,
            }
        )
    return pd.DataFrame(linhas)


def dataframe_toxic_relatorio(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(
            columns=["Data", "Nome", "Empresa", "CPF", "Status", "Enviado", "Cobrado"]
        )
    d = df.copy()
    d["Enviado"] = d["enviado"].map(_sim_nao_excel)
    d["Cobrado"] = d["cobrado"].map(_sim_nao_excel)
    d = d.rename(
        columns={
            "data_hora": "Data",
            "nome": "Nome",
            "empresa": "Empresa",
            "cpf": "CPF",
            "status_exame": "Status",
        }
    )
    return d[
        ["Data", "Nome", "Empresa", "CPF", "Status", "Enviado", "Cobrado"]
    ]


def montar_xlsx_toxic_com_totais(df_view: pd.DataFrame, titulo: str) -> bytes:
    """df_view já com colunas em português (últimas colunas Enviado/Cobrado texto)."""
    n = len(df_view)
    comp = int((df_view["Status"] == "COMPLETO").sum()) if n else 0
    parc = int((df_view["Status"] == "PARCIAL").sum()) if n else 0
    pend = int((df_view["Status"] == "PENDENTE").sum()) if n else 0

    linha_tot = {
        "Data": "",
        "Nome": f"{n} exames",
        "Empresa": "",
        "CPF": "",
        "Status": f"Comp:{comp} | Parc:{parc} | Pend:{pend}",
        "Enviado": "",
        "Cobrado": "",
    }
    linha_rot = {
        "Data": "TOTAIS:",
        "Nome": "",
        "Empresa": "",
        "CPF": "",
        "Status": "",
        "Enviado": "",
        "Cobrado": "",
    }

    df_out = pd.concat(
        [df_view, pd.DataFrame([linha_rot, linha_tot])],
        ignore_index=True,
    )

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Toxicológico", startrow=1, index=False)
        ws = writer.sheets["Toxicológico"]
        _formatar_planilha(ws, titulo=titulo, ncol=len(df_out.columns))
        mr = ws.max_row or 2
        for r in range(max(3, mr - 1), mr + 1):
            for col in range(1, len(df_out.columns) + 1):
                ws.cell(row=r, column=col).font = Font(bold=True)

    buf.seek(0)
    return buf.getvalue()


def excel_bytes_recepcao(df: pd.DataFrame, ref: date) -> bytes:
    titulo = f"RELATÓRIO DE RECEPÇÃO — {titulo_mes_ano(ref)}"
    v = dataframe_recepcao_relatorio(df)
    return montar_xlsx_generico(v, titulo, sheet_name="Recepção")


def excel_bytes_caixa(df: pd.DataFrame, ref: date) -> bytes:
    titulo = f"RELATÓRIO DE CAIXA — {titulo_mes_ano(ref)}"
    v = dataframe_caixa_relatorio(df)
    return montar_xlsx_generico(v, titulo, sheet_name="Caixa")


def excel_bytes_toxic(df: pd.DataFrame, ref: date) -> bytes:
    titulo = f"RELATÓRIO TOXICOLÓGICO — {titulo_mes_ano(ref)}"
    v = dataframe_toxic_relatorio(df)
    return montar_xlsx_toxic_com_totais(v, titulo)
